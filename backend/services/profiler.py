from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO, StringIO
from typing import Any

from openpyxl import load_workbook


MAX_SAMPLE_ROWS = 5


@dataclass(frozen=True)
class Profile:
    row_count: int
    column_count: int
    inferred_types: dict[str, str]
    missing_percent_by_column: dict[str, float]
    duplicate_rows: int
    sample_rows: list[dict[str, Any]]
    warnings: list[str]

    def to_dict(self) -> dict[str, Any]:
        return {
            "row_count": self.row_count,
            "column_count": self.column_count,
            "inferred_types": self.inferred_types,
            "missing_percent_by_column": self.missing_percent_by_column,
            "duplicate_rows": self.duplicate_rows,
            "sample_rows": self.sample_rows,
            "warnings": self.warnings,
        }


def profile_csv(content: bytes) -> Profile:
    text = _decode_csv(content)
    reader = csv.DictReader(StringIO(text))

    if not reader.fieldnames:
        raise ValueError("CSV must include a header row")

    columns = [_normalize_header(header) for header in reader.fieldnames]
    if any(not column for column in columns):
        raise ValueError("CSV headers must not be empty")
    if len(set(columns)) != len(columns):
        raise ValueError("CSV headers must be unique")

    rows: list[dict[str, str]] = []
    for raw_row in reader:
        row = {
            column: _clean_cell(raw_row.get(original_header, ""))
            for column, original_header in zip(columns, reader.fieldnames)
        }
        rows.append(row)

    return profile_rows(rows, columns, "CSV")


def profile_xlsx(content: bytes) -> Profile:
    if not content:
        raise ValueError("Uploaded file is empty")

    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("XLSX file could not be read") from exc

    worksheet = workbook.active
    row_iter = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(row_iter)
    except StopIteration as exc:
        raise ValueError("XLSX must include a header row") from exc

    columns = [_normalize_header(_cell_to_string(value)) for value in header_row]
    if any(not column for column in columns):
        raise ValueError("XLSX headers must not be empty")
    if len(set(columns)) != len(columns):
        raise ValueError("XLSX headers must be unique")

    rows: list[dict[str, str]] = []
    for raw_row in row_iter:
        values = list(raw_row[: len(columns)])
        values.extend([None] * (len(columns) - len(values)))
        row = {
            column: _clean_cell(_cell_to_string(value))
            for column, value in zip(columns, values)
        }
        if any(value != "" for value in row.values()):
            rows.append(row)

    return profile_rows(rows, columns, "XLSX")


def profile_rows(rows: list[dict[str, str]], columns: list[str], file_type: str) -> Profile:
    if not rows:
        raise ValueError(f"{file_type} must include at least one data row")

    duplicate_rows = _count_duplicate_rows(rows, columns)
    warnings = _build_warnings(rows, columns, duplicate_rows)

    return Profile(
        row_count=len(rows),
        column_count=len(columns),
        inferred_types={column: _infer_type(row[column] for row in rows) for column in columns},
        missing_percent_by_column={
            column: _missing_percent(row[column] for row in rows) for column in columns
        },
        duplicate_rows=duplicate_rows,
        sample_rows=rows[:MAX_SAMPLE_ROWS],
        warnings=warnings,
    )


def _decode_csv(content: bytes) -> str:
    if not content:
        raise ValueError("Uploaded file is empty")
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("CSV must be UTF-8 encoded") from exc
    if not text.strip():
        raise ValueError("Uploaded file is empty")
    return text


def _normalize_header(value: str | None) -> str:
    return (value or "").strip()


def _clean_cell(value: str | None) -> str:
    return (value or "").strip()


def _cell_to_string(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value)


def _count_duplicate_rows(rows: list[dict[str, str]], columns: list[str]) -> int:
    seen: set[tuple[str, ...]] = set()
    duplicates = 0
    for row in rows:
        row_key = tuple(row[column] for column in columns)
        if row_key in seen:
            duplicates += 1
        else:
            seen.add(row_key)
    return duplicates


def _missing_percent(values: Any) -> float:
    value_list = list(values)
    missing = sum(1 for value in value_list if value == "")
    return round((missing / len(value_list)) * 100, 2)


def _infer_type(values: Any) -> str:
    non_empty = [value for value in values if value != ""]
    if not non_empty:
        return "empty"
    if all(_is_integer(value) for value in non_empty):
        return "integer"
    if all(_is_number(value) for value in non_empty):
        return "number"
    if all(_is_date(value) for value in non_empty):
        return "date"
    if all(_is_boolean(value) for value in non_empty):
        return "boolean"
    return "string"


def _is_integer(value: str) -> bool:
    try:
        int(value)
    except ValueError:
        return False
    return "." not in value


def _is_number(value: str) -> bool:
    try:
        float(value)
    except ValueError:
        return False
    return True


def _is_date(value: str) -> bool:
    formats = ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S")
    return any(_matches_date_format(value, date_format) for date_format in formats)


def _matches_date_format(value: str, date_format: str) -> bool:
    try:
        datetime.strptime(value, date_format)
    except ValueError:
        return False
    return True


def _is_boolean(value: str) -> bool:
    return value.lower() in {"true", "false", "yes", "no", "0", "1"}


def _build_warnings(
    rows: list[dict[str, str]], columns: list[str], duplicate_rows: int
) -> list[str]:
    warnings: list[str] = []
    for column in columns:
        missing = _missing_percent(row[column] for row in rows)
        if missing > 0:
            warnings.append(f"{column} has {missing}% missing values")
    if duplicate_rows > 0:
        warnings.append(f"{duplicate_rows} duplicate row(s) detected")
    return warnings
