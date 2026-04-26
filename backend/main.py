import csv
import logging
import tempfile
from contextlib import asynccontextmanager
from io import BytesIO
from pathlib import Path

from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.responses import FileResponse, JSONResponse, PlainTextResponse
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook

from config import load_settings
from services.codegen import generate_python_script
from services.groundedness import check_and_revise
from services.planner import generate_analysis_plan
from services.profiler import profile_csv, profile_xlsx
from services.report_generator import generate_report, load_summary_json
from services.retry_runner import run_with_retry
from services.router import route_dataset
from services.run_store import RunStore, RunStatus
from services.sandbox_runner import (
    cleanup_expired_runs,
    create_run_directory,
    resolve_artifact_path,
)
from services.summary_validator import validate_summary_json

logger = logging.getLogger(__name__)

settings = load_settings()
run_store = RunStore()


@asynccontextmanager
async def lifespan(app: FastAPI):
    run_store.init_db()
    deleted = cleanup_expired_runs(settings.data_run_ttl_hours)
    if deleted:
        logger.info("startup: cleaned up %d expired run director(ies)", deleted)
    yield


app = FastAPI(title="DataBrief AI API", lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.cors_origins,
    allow_credentials=False,
    allow_methods=["GET", "POST"],
    allow_headers=["*"],
)


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok", "environment": settings.environment}


# ---------------------------------------------------------------------------
# Run status
# ---------------------------------------------------------------------------


@app.get("/api/runs/{run_id}")
def get_run_status(run_id: str) -> dict[str, object]:
    record = run_store.get(run_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Run not found.")
    return record.to_status_dict()


# ---------------------------------------------------------------------------
# Artifact serving
# ---------------------------------------------------------------------------


@app.get("/api/runs/{run_id}/artifacts/{artifact_name}")
def get_run_artifact(run_id: str, artifact_name: str) -> FileResponse:
    record = run_store.get(run_id)
    if record is not None and record.status == RunStatus.EXPIRED:
        raise HTTPException(status_code=410, detail="Run artifacts have expired.")
    artifact_path = resolve_artifact_path(run_id, artifact_name)
    if artifact_path is None:
        raise HTTPException(status_code=404, detail="Artifact not found.")
    return FileResponse(artifact_path)


# ---------------------------------------------------------------------------
# Export endpoints
# ---------------------------------------------------------------------------


@app.get("/api/runs/{run_id}/export/report.md")
def export_report_md(run_id: str) -> PlainTextResponse:
    record = run_store.get(run_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Run not found.")
    if record.status == RunStatus.EXPIRED:
        raise HTTPException(status_code=410, detail="Run artifacts have expired.")
    md = record.render_report_markdown()
    if md is None:
        raise HTTPException(status_code=404, detail="Report not available for this run.")
    return PlainTextResponse(md, media_type="text/markdown")


@app.get("/api/runs/{run_id}/export/findings.json")
def export_findings_json(run_id: str) -> JSONResponse:
    record = run_store.get(run_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Run not found.")
    if record.status == RunStatus.EXPIRED:
        raise HTTPException(status_code=410, detail="Run artifacts have expired.")
    findings = record.get_findings()
    if findings is None:
        raise HTTPException(status_code=404, detail="Findings not available for this run.")
    return JSONResponse(findings)


@app.get("/api/runs/{run_id}/export/analysis.py")
def export_analysis_py(run_id: str) -> PlainTextResponse:
    record = run_store.get(run_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Run not found.")
    if record.status == RunStatus.EXPIRED:
        raise HTTPException(status_code=410, detail="Run artifacts have expired.")
    code = record.generated_code
    if not code:
        raise HTTPException(status_code=404, detail="Generated code not available for this run.")
    return PlainTextResponse(code, media_type="text/x-python")


# ---------------------------------------------------------------------------
# Upload and analysis pipeline
# ---------------------------------------------------------------------------


@app.post("/api/upload")
async def upload_dataset(file: UploadFile = File(...)) -> dict[str, object]:
    if not file.filename:
        raise HTTPException(status_code=400, detail="Upload a CSV or XLSX file.")

    filename = file.filename.lower()
    if not filename.endswith((".csv", ".xlsx")):
        raise HTTPException(status_code=400, detail="Upload a CSV or XLSX file.")

    content = await file.read()
    max_bytes = settings.max_upload_mb * 1024 * 1024
    if len(content) > max_bytes:
        raise HTTPException(
            status_code=413,
            detail=f"File is larger than the {settings.max_upload_mb} MB limit.",
        )

    try:
        profile = profile_xlsx(content) if filename.endswith(".xlsx") else profile_csv(content)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    profile_payload = profile.to_dict()
    route = route_dataset(profile_payload)
    route_payload = route.to_dict()
    plan = generate_analysis_plan(profile_payload, route_payload)
    plan_payload = plan.to_dict()

    try:
        execution_input_path = _write_execution_input(content, filename)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    run_id, run_dir, artifact_dir = create_run_directory()

    # Persist initial run record so status endpoint works during execution.
    run_store.create(
        run_id=run_id,
        filename=file.filename,
        profile_json=profile_payload,
        route_json=route_payload,
        plan_json=plan_payload,
        ttl_hours=settings.data_run_ttl_hours,
    )

    generated_code = generate_python_script(
        profile=profile_payload,
        route=route_payload,
        plan=plan_payload,
        input_file_path=execution_input_path,
        artifact_dir=artifact_dir,
    )
    run_store.set_status(run_id, RunStatus.EXECUTING)
    run_store.set_generated_code(run_id, generated_code.code)

    # --- Bounded repair loop ---
    # input_file_path is kept alive until the loop completes.
    retry_result = run_with_retry(
        code=generated_code.code,
        artifact_dir=artifact_dir,
        profile=profile_payload,
        route=route_payload,
        plan=plan_payload,
        input_file_path=execution_input_path,
    )

    try:
        execution_input_path.unlink(missing_ok=True)
        execution_input_path.parent.rmdir()
    except OSError:
        pass

    run_store.set_status(run_id, RunStatus.EVALUATING)

    # Validate summary.json if present.
    summary_data = load_summary_json(retry_result.final_execution.artifacts)
    summary_errors = validate_summary_json(summary_data) if summary_data else [
        "summary.json was not produced by execution"
    ]
    if summary_errors:
        logger.warning("run %s: summary.json validation errors: %s", run_id, summary_errors)

    computed_facts = {
        **summary_data,
        "row_count": profile_payload.get("row_count", 0),
        "column_count": profile_payload.get("column_count", 0),
        "duplicate_rows": profile_payload.get("duplicate_rows", 0),
        "dataset_type": route_payload.get("dataset_type", "generic"),
    }

    report = generate_report(
        profile=profile_payload,
        route=route_payload,
        plan=plan_payload,
        evaluation=retry_result.final_evaluation,
        execution=retry_result.final_execution,
    )

    # Single-pass groundedness revision.
    report = check_and_revise(report, computed_facts)

    final_status = (
        RunStatus.COMPLETE
        if retry_result.final_evaluation.outcome == "success"
        else RunStatus.PARTIAL
        if retry_result.final_execution.artifacts
        else RunStatus.FAILED
    )
    run_store.finish(
        run_id=run_id,
        status=final_status,
        retry_count=retry_result.retry_count,
        evaluation_result=retry_result.final_evaluation.to_dict(),
        report_json=report.to_dict(),
        summary_errors=summary_errors,
    )

    return {
        "run_id": run_id,
        "filename": file.filename,
        "profile": profile_payload,
        "route": route_payload,
        "plan": plan_payload,
        "codegen": generated_code.to_dict(),
        "execution": retry_result.final_execution.to_dict(),
        "retry": retry_result.to_dict(),
        "report": report.to_dict(),
        "summary_validation": {"errors": summary_errors},
    }


def _write_execution_input(content: bytes, filename: str) -> Path:
    temp_dir = Path(tempfile.mkdtemp(prefix="databrief-ai-input-"))
    if filename.endswith(".csv"):
        csv_path = temp_dir / "upload.csv"
        csv_path.write_bytes(content)
        return csv_path

    csv_path = temp_dir / "upload.csv"
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("XLSX file could not be prepared for execution") from exc

    worksheet = workbook.active
    with csv_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        for row in worksheet.iter_rows(values_only=True):
            writer.writerow(["" if value is None else value for value in row])

    return csv_path
